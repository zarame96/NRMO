# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
rec=json.load(open("/tmp/listed_record.json"))
A=rec["annual"]
YEN=1_000_000  # 内部は百万円→円
def y(v): return round(float(v)*YEN)

BLUE=Font(name="Arial",color="0000FF",size=10); BLACK=Font(name="Arial",color="000000",size=10)
BOLD=Font(name="Arial",bold=True,size=10,color="000000"); HEAD=Font(name="Arial",bold=True,color="FFFFFF",size=10)
NAVY=PatternFill("solid",fgColor="1B2748"); GREY=PatternFill("solid",fgColor="EEF1F6")
TITLE=Font(name="Arial",bold=True,size=13,color="1B2748"); ITAL=Font(name="Arial",italic=True,size=9,color="555555")
MONEY='#,##0;(#,##0);"-"'
def setcol(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
def hdr(ws,row,labels,wrap=True):
    for j,l in enumerate(labels,1):
        c=ws.cell(row,j,l); c.font=HEAD; c.fill=NAVY; c.alignment=Alignment(horizontal="center",wrap_text=wrap,vertical="center")

wb=Workbook(); wb.remove(wb.active)

# ===== 表紙 =====
ws=wb.create_sheet("表紙")
setcol(ws,[26,80])
ws["A1"]="100年経営シミュレーション — 年次連結財務諸表(IFRS, 円)"; ws["A1"].font=TITLE
rows=[
 ("基準","国際会計基準(IFRS)に準拠した表示。単位: 円。"),
 ("作成","本物の NRMO(MaxForwardEngine)を無改変で駆動した最大100年の経営シミュレーション。複式簿記で資産=負債+純資産を全期保証。"),
 ("企業の性質",f"未上場で創業 → 第{[a for a in A if a['bs']['share']>A[0]['bs']['share']*1.5][0]['year'] if any(a['bs']['share']>A[0]['bs']['share']*1.5 for a in A) else '-'}期前後にIPO(上場) → 増資・公募で資本増強。"),
 ("存続",f"{rec['lifespan']:.0f}年 / 第{rec['generations']}代まで経営者交代 / 危機(立て直し局面){len(rec['crises'])}回 / {'継続中(100年到達)' if not rec['bankrupt'] else '破綻: '+rec['death_reason']}"),
 ("規模",f"資本金 {y(A[0]['bs']['share']):,}円 → {y(A[-1]['bs']['share']):,}円 / 純資産 {y(A[0]['bs']['equity']):,}円 → {y(A[-1]['bs']['equity']):,}円"),
 ("判断軸","経営者は世代ごとに交代し、リスク選好(保守〜攻撃)が揺れる。NRMOの破滅回避の核は全世代で不変。"),
 ("継続企業の前提","各期の純資産がマイナスとなった場合『継続企業の前提に重要な疑義』を注記。資金繰りが融資・増資のいずれでも埋められない場合に限り破綻。"),
 ("重要な注意","数値は研究用プロキシであり実在企業の実額ではない。実データによる較正は未実施。経営判断の根拠に用いないこと。"),
]
r=3
for k,v in rows:
    ws.cell(r,1,k).font=BOLD; c=ws.cell(r,2,v); c.alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=30; r+=1

# ===== 損益計算書(P/L) =====
ws=wb.create_sheet("損益計算書(P_L)")
setcol(ws,[7]+[15]*11)
ws["A1"]="損益計算書 (IFRS, 円) — 年次"; ws["A1"].font=TITLE
ws["A2"]="売上総利益・営業利益・税引前利益・当期純利益は数式。減損損失は当期純利益の下で純資産から控除(包括的)。"; ws["A2"].font=ITAL
cols=["年度","売上高","売上原価","売上総利益","販管費","減価償却費","営業利益","支払利息","税引前利益","法人税等","当期純利益","減損損失"]
hdr(ws,4,cols)
for i,a in enumerate(A):
    R=5+i; pl=a["pl"]
    ws.cell(R,1,str(a["year"])).font=BLACK; ws.cell(R,1).alignment=Alignment(horizontal="center")
    vals={2:pl["revenue"],3:pl["cogs"],5:pl["opex"],6:pl["dep"],8:pl["interest"],10:pl["tax"],12:pl.get("impairment",0)}
    for col,v in vals.items(): c=ws.cell(R,col,y(v)); c.font=BLUE; c.number_format=MONEY
    cl=get_column_letter
    ws.cell(R,4,f"={cl(2)}{R}-{cl(3)}{R}").font=BOLD              # 売上総利益
    ws.cell(R,7,f"={cl(4)}{R}-{cl(5)}{R}-{cl(6)}{R}").font=BOLD   # 営業利益
    ws.cell(R,9,f"={cl(7)}{R}-{cl(8)}{R}")                         # 税引前
    ws.cell(R,11,f"={cl(9)}{R}-{cl(10)}{R}").font=BOLD            # 当期純利益
    for col in (4,7,9,11): ws.cell(R,col).number_format=MONEY
ws.freeze_panes="B5"

# ===== 貸借対照表(B/S) =====
ws=wb.create_sheet("貸借対照表(B_S)")
setcol(ws,[7]+[13]*14)
ws["A1"]="貸借対照表 (IFRS, 円) — 年次 / 流動・非流動区分"; ws["A1"].font=TITLE
ws["A2"]="合計・純資産・検証(資産-(負債+純資産)=0)は数式。"; ws["A2"].font=ITAL
cols=["年度","現金","売掛金","棚卸資産","流動資産計","純固定資産","資産合計","買掛金","短期借入","流動負債計","長期借入","負債合計","資本金","利益剰余金","純資産計","検証"]
# 16列に拡張
setcol(ws,[7]+[12]*15)
hdr(ws,4,cols)
for i,a in enumerate(A):
    R=5+i; bs=a["bs"]; cl=get_column_letter
    ws.cell(R,1,str(a["year"])).font=BLACK; ws.cell(R,1).alignment=Alignment(horizontal="center")
    inp={2:bs["cash"],3:bs["ar"],4:bs["inv"],6:bs["net_ppe"],8:bs["tp"],9:bs["bc"],11:bs["bnc"],13:bs["share"],14:bs["retained"]}
    for col,v in inp.items(): c=ws.cell(R,col,y(v)); c.font=BLUE; c.number_format=MONEY
    ws.cell(R,5,f"={cl(2)}{R}+{cl(3)}{R}+{cl(4)}{R}").font=BOLD          # 流動資産計
    ws.cell(R,7,f"={cl(5)}{R}+{cl(6)}{R}").font=BOLD                     # 資産合計
    ws.cell(R,10,f"={cl(8)}{R}+{cl(9)}{R}").font=BOLD                    # 流動負債計
    ws.cell(R,12,f"={cl(10)}{R}+{cl(11)}{R}").font=BOLD                  # 負債合計
    ws.cell(R,15,f"={cl(13)}{R}+{cl(14)}{R}").font=BOLD                  # 純資産計
    ws.cell(R,16,f"={cl(7)}{R}-({cl(12)}{R}+{cl(15)}{R})")               # 検証=0
    for col in (5,7,10,12,15,16): ws.cell(R,col).number_format=MONEY
    if a.get("going_concern_doubt"):
        ws.cell(R,1).fill=PatternFill("solid",fgColor="F8D7DA")          # 継続企業の疑義=赤
ws.freeze_panes="B5"

# ===== キャッシュフロー計算書(C/F) =====
ws=wb.create_sheet("CF計算書")
setcol(ws,[7,15,15,15,16,15,15])
ws["A1"]="キャッシュ・フロー計算書 (IFRS, 円, 間接法) — 年次"; ws["A1"].font=TITLE
ws["A2"]="現金増減=営業+投資+財務(数式)。財務CFのうち増資・IPOによる調達額をmemo表示。"; ws["A2"].font=ITAL
cols=["年度","営業CF","投資CF","財務CF","現金増減","(memo)増資・IPO","(memo)新規借入"]
hdr(ws,4,cols)
for i,a in enumerate(A):
    R=5+i; pl=a["pl"]; cl=get_column_letter
    ws.cell(R,1,str(a["year"])).font=BLACK; ws.cell(R,1).alignment=Alignment(horizontal="center")
    for col,v in {2:pl["cfo"],3:pl["cfi"],4:pl["cff"],6:pl.get("new_equity",0),7:pl.get("new_debt",0)}.items():
        c=ws.cell(R,col,y(v)); c.font=BLUE; c.number_format=MONEY
    ws.cell(R,5,f"={cl(2)}{R}+{cl(3)}{R}+{cl(4)}{R}").font=BOLD
    ws.cell(R,5).number_format=MONEY
ws.freeze_panes="B5"

# ===== 経営史(世代別) =====
ws=wb.create_sheet("経営史(世代別)")
setcol(ws,[6,14,9,11,14,8,10,22])
ws["A1"]="経営史 — 世代別の判断軸と帰結"; ws["A1"].font=TITLE
ws["A2"]="経営者は交代し判断軸(保守〜攻撃)が揺れる。NRMOの破滅回避は不変。"; ws["A2"].font=ITAL
cols=["代","判断軸","軸の値","在任(年)","期末純資産(円)","危機回数","主な手","経験レジーム"]
hdr(ws,4,cols)
def blab(b): return "攻撃的" if b>0.33 else ("保守的" if b<-0.33 else "均衡")
r=5
for e in rec["eras"]:
    if e.get("end_year",0)-e["start_year"]<1 and e["gen"]>1: continue
    am=max(e["actions"],key=e["actions"].get) if e["actions"] else "-"
    am_jp={"invest":"設備投資","explore":"販促/刷新","equity":"増資","ipo":"上場","defend":"価格防衛","recover":"資金確保","finance":"借入","divest":"事業売却"}.get(am,am)
    row=[e["gen"],blab(e["bias"]),round(e["bias"],2),f"{e['start_year']:.0f}-{e.get('end_year',0):.0f}",y(e.get("end_equity",0)),e["crises"],am_jp,"/".join(e["regimes"])]
    for j,v in enumerate(row,1):
        c=ws.cell(r,j,v); c.font=BLACK
        if j==5: c.number_format=MONEY
    r+=1

wb.save("/mnt/user-data/outputs/NRMO_100年_IFRS財務三表.xlsx")
print("xlsx 保存")
