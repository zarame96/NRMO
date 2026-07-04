# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C=json.load(open("/tmp/century_full.json",encoding="utf-8"))
Y=json.load(open("/tmp/young_full.json",encoding="utf-8"))

def _norm(r):
    r["gens"]=r.get("generations", r.get("gens",1))
    r["peak_eq"]=r.get("peak_equity",0)/1e6
    r["death"]=r.get("death_reason","破綻") or "存続中"
    for e in r.get("eras",[]):
        acts=e.get("actions",{}); e["top_action"]=max(acts,key=acts.get) if acts else "-"
    return r
C=_norm(C); Y=_norm(Y)

BLUE=Font(name="Arial",color="0000FF",size=10); BLACK=Font(name="Arial",color="000000",size=10)
HEAD=Font(name="Arial",bold=True,color="FFFFFF",size=10); TITLE=Font(name="Arial",bold=True,size=13,color="1B2748")
SUB=Font(name="Arial",size=9,italic=True,color="555555"); SECT=Font(name="Arial",bold=True,size=10,color="1B2748")
NAVY=PatternFill("solid",fgColor="1B2748"); GREY=PatternFill("solid",fgColor="EEF1F6")
MONEY='#,##0;(#,##0);"-"'; thin=Side(style="thin",color="D7DCE5")

def bias_jp(b): return "攻撃" if b>0.33 else ("保守" if b<-0.33 else "均衡")

def stmt_sheet(wb, rec, name, title, note):
    ws=wb.create_sheet(name); ann=rec["annual"]
    ws["A1"]=title; ws["A1"].font=TITLE
    ws["A2"]=note; ws["A2"].font=SUB
    # P/L → B/S → C/F を年=行で3ブロック横並びにせず、3シートに分けると重いので1シートに区切る
    # 列: 年, 世代, 構え, 景気, [P/L...], [B/S...], [C/F...]
    cols=[("年","year",None),("代","gen",None),("構え","bias_jp",None),("景気","regime",None),
          # P/L (J-GAAP)
          ("売上高","pl.revenue","i"),("売上原価","pl.cogs","i"),("売上総利益",None,"=rev-cogs"),
          ("販管費","pl.opex","i"),("減価償却費","pl.dep","i"),("営業利益",None,"=gp-opex-dep"),
          ("支払利息","pl.interest","i"),("経常利益",None,"=op-int"),
          ("減損損失(特別)","pl.impairment","i"),("税引前純利益",None,"=ord-imp"),
          ("法人税等","pl.tax","i"),("当期純利益",None,"=pre-tax"),
          # B/S
          ("現金","bs.cash","i"),("売掛金","bs.ar","i"),("棚卸資産","bs.inv","i"),("有形固定資産(純)","bs.net_ppe","i"),
          ("資産合計",None,"=cash+ar+inv+ppe"),
          ("買掛金","bs.tp","i"),("短期借入","bs.bc","i"),("長期借入","bs.bnc","i"),("負債合計",None,"=tp+bc+bnc"),
          ("資本金","bs.share","i"),("利益剰余金","bs.retained","i"),("純資産合計",None,"=share+ret"),
          ("検証(資産-(負債+純資産))",None,"=check"),
          # C/F
          ("営業CF","pl.cfo","i"),("投資CF","pl.cfi","i"),("財務CF","pl.cff","i"),("現金増減",None,"=cfo+cfi+cff"),
          ("継続企業の前提疑義","gc","g")]
    hr=4
    for j,(label,_,_) in enumerate(cols,1):
        c=ws.cell(hr,j,label); c.font=HEAD; c.fill=NAVY; c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
    ws.freeze_panes="E5"
    # 列インデックス辞書(数式用)
    idx={label:get_column_letter(j) for j,(label,_,_) in enumerate(cols,1)}
    for i,a in enumerate(ann):
        r=hr+1+i
        def val(path):
            o=a
            try:
                for p in path.split("."): o=o[p]
            except KeyError:
                return 0.0
            return o
        for j,(label,src,kind) in enumerate(cols,1):
            cell=ws.cell(r,j); col=get_column_letter(j)
            if src=="year": cell.value=a["year"]; cell.font=BLACK; cell.alignment=Alignment(horizontal="center")
            elif src=="gen": cell.value=a["gen"]; cell.font=BLACK; cell.alignment=Alignment(horizontal="center")
            elif src=="bias_jp": cell.value=bias_jp(a["bias"]); cell.font=BLACK; cell.alignment=Alignment(horizontal="center")
            elif src=="regime": cell.value=a["regime"]; cell.font=BLACK; cell.alignment=Alignment(horizontal="center")
            elif src=="gc": cell.value="疑義" if a["going_concern_doubt"] else ""; cell.font=Font(name="Arial",size=9,color="C0392B" if a["going_concern_doubt"] else "000000")
            elif kind=="i":
                cell.value=round(val(src),1); cell.font=BLUE; cell.number_format=MONEY
            else:  # 数式(小計)
                R=r; f={
                  "=rev-cogs":f'={idx["売上高"]}{R}-{idx["売上原価"]}{R}',
                  "=gp-opex-dep":f'={idx["売上総利益"]}{R}-{idx["販管費"]}{R}-{idx["減価償却費"]}{R}',
                  "=op-int":f'={idx["営業利益"]}{R}-{idx["支払利息"]}{R}',
                  "=ord-imp":f'={idx["経常利益"]}{R}-{idx["減損損失(特別)"]}{R}',
                  "=pre-tax":f'={idx["税引前純利益"]}{R}-{idx["法人税等"]}{R}',
                  "=cash+ar+inv+ppe":f'={idx["現金"]}{R}+{idx["売掛金"]}{R}+{idx["棚卸資産"]}{R}+{idx["有形固定資産(純)"]}{R}',
                  "=tp+bc+bnc":f'={idx["買掛金"]}{R}+{idx["短期借入"]}{R}+{idx["長期借入"]}{R}',
                  "=share+ret":f'={idx["資本金"]}{R}+{idx["利益剰余金"]}{R}',
                  "=check":f'={idx["資産合計"]}{R}-({idx["負債合計"]}{R}+{idx["純資産合計"]}{R})',
                  "=cfo+cfi+cff":f'={idx["営業CF"]}{R}+{idx["投資CF"]}{R}+{idx["財務CF"]}{R}',
                }[kind]
                cell.value=f; cell.font=BLACK; cell.number_format=MONEY
    # 列幅
    ws.column_dimensions["A"].width=5; ws.column_dimensions["B"].width=4; ws.column_dimensions["C"].width=5; ws.column_dimensions["D"].width=7
    for j in range(5,len(cols)+1): ws.column_dimensions[get_column_letter(j)].width=11

def history_sheet(wb, rec):
    ws=wb.create_sheet("経営史(世代)"); 
    ws["A1"]="100年(最大)経営史 — 世代ごとの判断軸と帰結"; ws["A1"].font=TITLE
    ws["A2"]=f"代表企業(seed{rec['seed']}) / {rec['lifespan']:.0f}年生存 / 第{rec['gens']}代 / {rec['market'].upper()}上場・{rec['accounting'].upper()}適用 / 最終監査={rec['audit']} / ピーク純資産{rec['peak_eq']:,.0f}百万円"; ws["A2"].font=SUB
    hdr=["代","判断軸(bias)","在任(年)","期末純資産(百万円)","危機回数","経験した景気","主な手","制度イベント"]
    ws.append([]); ws.append(hdr); hr=ws.max_row
    for j,h in enumerate(hdr,1): c=ws.cell(hr,j); c.font=HEAD; c.fill=NAVY; c.alignment=Alignment(wrap_text=True,horizontal="center",vertical="center")
    AMAP={"invest":"設備投資","explore":"販促/刷新","defend":"価格防衛","recover":"資金確保/返済","finance":"借入","divest":"事業売却","equity":"増資","ipo":"上場","hold":"様子見","newbiz":"新規事業","trial":"試行R&D","acquire":"集客/広告","channel":"多チャネル","openstore":"出店","insource":"内製化","cs":"CS強化","optimize":"事業最適化","pivot":"業態転換","bond":"社債発行","convbond":"CB発行","dividend":"配当","buyback":"自社株買い"}
    for e in rec["eras"]:
        if e.get("end_year",0)-e["start_year"]<1 and e["gen"]>1: continue
        ws.append([f"第{e['gen']}代", f"{bias_jp(e['bias'])} ({e['bias']:+.2f})",
                   f"{e['start_year']:.0f}→{e.get('end_year',0):.0f}", round(e.get("end_equity",0)),
                   e["crises"], "/".join(e["regimes"]), AMAP.get(e["top_action"],e["top_action"]),
                   "; ".join(e["events"]) if e["events"] else "—"])
        r=ws.max_row; ws.cell(r,4).number_format=MONEY; ws.cell(r,4).font=BLACK
    for col,w in {"A":6,"B":13,"C":11,"D":18,"E":8,"F":18,"G":12,"H":40}.items(): ws.column_dimensions[col].width=w
    # 危機(立て直し)明細
    ws.append([]); ws.append(["危機(立て直し)局面の明細"]); ws.cell(ws.max_row,1).font=SECT
    ws.append(["代","開始年","終了年","結果","底の純資産(百万円)"])
    r=ws.max_row
    for j in range(1,6): ws.cell(r,j).font=HEAD; ws.cell(r,j).fill=GREY; ws.cell(r,j).font=SECT
    for c in rec["crises"]:
        ws.append([f"第{c['gen']}代", round(c["start"],1), round(c["end"],1), "生還" if c["survived"] else "破綻", c["low_equity"]])


def market_sheet(wb, rec, name, title):
    ws=wb.create_sheet(name); ann=rec["annual"]
    ws["A1"]=title; ws["A1"].font=TITLE
    ws["A2"]="単位：シェア=%, 市場規模=個, 株価=円, 金額=百万円 / 値はシミュレーション値"; ws["A2"].font=SUB
    ws["A3"]=f"企業理念: 『{rec.get('creed','')}』  最終 理念整合度={rec.get('creed_align',1.0):.2f}"; ws["A3"].font=SECT
    cols=["年","世代","市場シェア","競合数","市場規模(個)","新規事業(累計)","上場","市場","会計","監査意見","株価(円)","理念整合度","アクティビスト圧","配当累計(百万円)","社債残(百万円)","自己株式(百万円)","発行済株式(百万株)","チャネル数","拠点数","内製化","CS水準","opex効率","業態転換(累計)"]
    hr=5
    for j,h in enumerate(cols,1):
        c=ws.cell(hr,j,h); c.font=HEAD; c.fill=NAVY; c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
    ws.freeze_panes="C6"
    for i,a in enumerate(ann):
        r=hr+1+i
        row=[a["year"],a["gen"],a.get("market_share",0),a.get("n_competitors",0),int(a.get("market_size",0)),
             a.get("new_biz",0),"上場" if a.get("listed") else "未上場",a.get("market","-").upper(),a.get("accounting","-").upper(),
             a.get("audit","-"),round(a.get("share_price",0)),round(a.get("creed_align",1.0),2),a.get("activist",0),
             round(a.get("dividends_cum",0),1),round(a.get("bonds",0),1),round(a.get("treasury",0),1),round(a.get("shares",100),1),
             a.get("channels",1),a.get("sites",1),a.get("insourcing",0.0),a.get("cs_level",0.0),a.get("opex_efficiency",1.0),a.get("pivot_count",0)]
        for j,v in enumerate(row,1):
            c=ws.cell(r,j,v); c.font=BLACK
            if j==3: c.number_format='0.0%'
            if j==11: c.number_format=MONEY     # 株価(円)
            if j in (14,15,16): c.number_format=MONEY
            if j in (20,21,22): c.number_format='0.00'   # 内製化/CS/opex効率
            if j in (1,2,4,6,18,19,23): c.alignment=Alignment(horizontal="center")
    widths={"A":5,"B":5,"C":10,"D":7,"E":13,"F":13,"G":8,"H":7,"I":8,"J":13,"K":12,"L":11,"M":13,"N":15,"O":14,"P":14,"Q":16,"R":9,"S":7,"T":8,"U":8,"V":9,"W":13}
    for col,w in widths.items(): ws.column_dimensions[col].width=w

wb=Workbook(); wb.remove(wb.active)
history_sheet(wb, C)
market_sheet(wb, C, "代表-市場と資本政策", f"代表企業 市場シェア・競合・新規事業・資本政策・操業レバーの推移（{C['lifespan']:.0f}年 / {C['accounting'].upper()}・{C['market'].upper()}）")
stmt_sheet(wb, C, "代表-財務三表",
           f"代表企業 {C['lifespan']:.0f}年 財務三表（{C['accounting'].upper()}・{C['market'].upper()}上場 / 最大100年モデル）",
           "単位：百万円 / 行=年, 列=科目 / 青=入力, 黒=数式(小計) / 検証列=資産-(負債+純資産)=0 / 数値はシミュレーション値(実額較正前)")
stmt_sheet(wb, Y, f"対照-零細({Y['lifespan']:.0f}年で破綻)",
           f"対照: 未上場の零細企業（{Y['lifespan']:.0f}年で{Y['death']}）",
           "単位：百万円 / 上場せず資本市場アクセスなく流動性破綻 / 零細が潰れるのは当たり前")
wb.save("/mnt/user-data/outputs/NRMO_100年_経営史と財務三表.xlsx")
print("xlsx 作成")
